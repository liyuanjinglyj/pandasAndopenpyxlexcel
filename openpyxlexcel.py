import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

'''
wb=openpyxl.load_workbook('1234.xlsx')
sheet=wb.worksheets[0]
sheet.cell(1,1,'111111')
font = Font(name='微软雅黑', size=10, bold=False, italic=False, vertAlign=None,
                                        underline='none', strike=False, color='FFFF0000')
sheet.cell(1,1).font=font
fill = PatternFill(fill_type='darkUp',start_color='FFFF00',end_color='FF0000')
sheet.cell(1,1).fill=fill

border = Border(left=Side(border_style='dashDotDot',color='9932CC'),
                right=Side(border_style='dashDotDot',color='121212'),
                top=Side(border_style='dashDotDot',color='8B0A50'),
                bottom=Side(border_style='dashDotDot',color='B3EE3A'),)
sheet.cell(5,4).border=border
alignment = Alignment(horizontal='center',
                    vertical='center',
                    text_rotation=0,
                    indent=22)

sheet.cell(5,3).alignment=alignment
sheet.cell(6,3).value='=HYPERLINK("%s","%s")' % ("#A1", "跳转到A1")
wb.save('1234.xlsx')'''
wb = openpyxl.load_workbook('sum.xlsx')
sheet = wb.worksheets[0]
sheet['F1']='=SUM(A1:E1)'

sheet.unmerge_cells('A1:E1')
wb.save('sum.xlsx')